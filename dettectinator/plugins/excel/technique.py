"""
Dettectinator - The Python library to your DeTT&CT YAML files.
Authors:
    Martijn Veken, Sirius Security
    Ruben Bouman, Sirius Security
License: GPL-3.0 License
"""

from argparse import ArgumentParser
from collections.abc import Iterable
from plugins.base.technique import TechniqueBase


class TechniqueExcel(TechniqueBase):
    """
    Import data from an Excel file, having a worksheet with two columns: TechniqueId and UseCase
    """

    def __init__(self, parameters: dict) -> None:
        super().__init__(parameters)
        if 'file' not in self._parameters:
            raise Exception('DetectionExcel: "file" parameter is required.')

    @staticmethod
    def set_plugin_params(parser: ArgumentParser) -> None:
        """
        Set command line arguments specific for the plugin
        :param parser: Argument parser
        """
        TechniqueBase.set_plugin_params(parser)

        parser.add_argument('--file', help='Path of the Excel file to import', required=True)

    def get_data_from_source(self) -> Iterable:
        """
        Gets the use-case/technique data from the source.
        :return: Iterable, yields technique, detection
        """
        file = self._parameters['file']
        print(f'Reading data from "{file}"')

        import openpyxl
        wb = openpyxl.load_workbook(filename=file, data_only=True)
        sheet = wb.worksheets[0]

        for rowNumber in range(2, sheet.max_row + 1):
            techniques = sheet.cell(row=rowNumber, column=1).value
            detection = sheet.cell(row=rowNumber, column=2).value
            for technique in techniques.split(','):
                yield technique.strip(), detection
